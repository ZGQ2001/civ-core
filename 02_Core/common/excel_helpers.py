"""Excel（pandas / openpyxl）通用操作。

读列表用 pandas；改单元格保留格式时用 openpyxl 直改 —— 不要 pandas + to_excel，
那样会丢失原工作簿的样式 / 合并 / 公式。
"""

from openpyxl import load_workbook

from common.patterns import FIG_PATTERN
from common.word_helpers import make_caption_substitutor


def get_excel_sort_order(
    excel_path: str, col_name: str, sheet_name: str | None = None
) -> list[int]:
    """读 Excel 指定列，把每行的"图 N"提成数字列表，作为排序顺序。"""
    import pandas as pd

    try:
        df = (
            pd.read_excel(excel_path, sheet_name=sheet_name)
            if sheet_name
            else pd.read_excel(excel_path)
        )
        if col_name not in df.columns:
            print(f"❌ Sheet [{sheet_name or '默认'}] 中找不到列：{col_name}")
            return []
        order: list[int] = []
        for item in df[col_name].dropna().astype(str).tolist():
            m = FIG_PATTERN.search(item)
            if m:
                order.append(int(m.group(1)))
        return order
    except Exception as e:
        print(f"❌ 读取 Excel 失败: {e}")
        return []


def find_column_index(ws, col_name: str, header_row: int = 1) -> int | None:
    """在指定 sheet 的表头行里找列名对应的列号（1-indexed），找不到返回 None。"""
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip() == col_name:
            return cell.column
    return None


def replace_in_excel_column(
    excel_path: str,
    sheet_name: str | None,
    col_name: str,
    mapping: dict,
    output_path: str,
    header_row: int = 1,
) -> tuple[int, list[int]]:
    """用 openpyxl 直改指定列的"图 N"，保留所有其他格式。

    返回 (改动了的单元格数, 找不到映射的旧编号列表)。
    """
    wb = load_workbook(excel_path)
    target = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target]

    col_idx = find_column_index(ws, col_name, header_row)
    if col_idx is None:
        raise ValueError(f"❌ Sheet [{target}] 中找不到列：{col_name}")

    apply, unmatched = make_caption_substitutor(mapping)

    replaced = 0
    for row in ws.iter_rows(min_row=header_row + 1, min_col=col_idx, max_col=col_idx):
        cell = row[0]
        if cell.value is None:
            continue
        s = str(cell.value)
        if not FIG_PATTERN.search(s):
            continue
        new_s = apply(s)
        if new_s != s:
            cell.value = new_s
            replaced += 1

    wb.save(output_path)
    return replaced, unmatched
