"""Excel 输入读取与清洗（openpyxl-only，禁止 pandas）。

为什么不用 pandas：
  • CLAUDE.md 总纲明令"Excel openpyxl，禁用 pandas"
  • pandas 对 NaN 的隐式转换会把空单元格变成 float('nan')，业务侧每次都得加 isna 判断
  • 只读模式下 openpyxl 内存峰值远低于 pandas，仪器导出几万行也稳

对外暴露三个函数：
  • read_sheet_names(path)                          → list[str]
  • read_rows(path, sheet_name, header_row)         → list[dict[str, Any]]
  • get_column_headers(path, sheet_name, header_row)→ list[str]

清洗策略（一次到位，下游不用再处理）：
  • 表头两端空白全部 strip；全空白/None 的表头列直接丢弃
  • 整行全空（None 或空白字符串）的数据行整行跳过
  • 单元格值保留 openpyxl 的原生类型（int/float/str/datetime/None），
    数值转换由调用方按业务语义决定
"""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from civ_core.utils.logger import get_logger

log = get_logger(__name__)


# ──────────────────────────────────────────────────────────────────
# 异常
# ──────────────────────────────────────────────────────────────────
class ExcelReadError(RuntimeError):
    """读取 Excel 失败。携带 hint 字段供 UI 三段式提示直接展示。"""

    hint: str

    def __init__(self, message: str, *, hint: str = "") -> None:
        super().__init__(message)
        self.hint = hint


# ──────────────────────────────────────────────────────────────────
# 内部：workbook 资源管理
# ──────────────────────────────────────────────────────────────────
@contextmanager
def _opened(path: Path) -> Iterator[Workbook]:
    """以只读模式打开 Workbook，并在 finally 中确保 close。

    read_only=True 下 openpyxl 走流式读取，内存友好；
    data_only=True 让公式单元格返回最后一次缓存的计算值（与肉眼所见一致）。
    """
    if not path.is_file():
        raise ExcelReadError(
            f"Excel 文件不存在：{path}",
            hint=f"请检查路径是否正确：{path}",
        )
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelReadError(
            f"无法打开 Excel（{path.name}）：{e}",
            hint="可能是文件损坏、被其他程序锁定或不是合法的 xlsx/xlsm。",
        ) from e
    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception as e:
            # close 失败不应影响主流程；仅记录
            log.warning("关闭 workbook 时出错（已忽略）：%s", e)


def _resolve_sheet(wb: Workbook, sheet_name: str | None) -> str:
    """统一的 sheet 解析：None → 第一个；不存在 → 抛带 hint 的异常。"""
    if sheet_name is None:
        return wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ExcelReadError(
            f"Sheet 不存在：{sheet_name!r}",
            hint=f"该 Excel 实际可用 sheet：{wb.sheetnames}",
        )
    return sheet_name


def _is_blank(value: Any) -> bool:
    """判断单元格是否视为"空"：None 或全空白字符串。"""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


# ──────────────────────────────────────────────────────────────────
# 公共 API
# ──────────────────────────────────────────────────────────────────
def read_sheet_names(excel_path: Path | str) -> list[str]:
    """返回 Excel 所有 sheet 名（按文件顺序）。失败抛 ExcelReadError。"""
    path = Path(excel_path)
    with _opened(path) as wb:
        return list(wb.sheetnames)


def get_column_headers(
    excel_path: Path | str,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> list[str]:
    """快速读取指定 sheet 的表头（不加载数据），用于 UI 列名校验/预览。

    仅返回非空表头列，并 strip 两端空白；保持 Excel 列顺序。
    """
    if header_row < 1:
        raise ExcelReadError(f"header_row 必须 >= 1，得到 {header_row}")

    path = Path(excel_path)
    with _opened(path) as wb:
        sheet = _resolve_sheet(wb, sheet_name)
        ws = wb[sheet]
        # 只读 header_row 一行；read_only 模式下 iter_rows 是生成器，提前 break 不会全表扫
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == header_row:
                return [str(c).strip() for c in row if not _is_blank(c)]
            if i > header_row:
                break
    return []


def read_rows(
    excel_path: Path | str,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> list[dict[str, Any]]:
    """读取一个 sheet 的全部数据，每行返回 {表头列名: 单元格值} 字典。

    设计要点：
      • 表头从 `header_row`（1-based）读取，列名 strip 后作为 dict key
      • 全空白/None 的表头列被丢弃，对应数据列也不会出现在返回的 dict 里
      • 整行全空的数据行整行跳过（仪器导出 xlsx 末尾常有大量空白行）
      • 单元格值保留 openpyxl 原生类型，由调用方按业务语义转 float/str

    异常：
      • ExcelReadError：文件不存在 / 打不开 / sheet 不存在 / header_row 非法
    """
    if header_row < 1:
        raise ExcelReadError(f"header_row 必须 >= 1，得到 {header_row}")

    path = Path(excel_path)
    with _opened(path) as wb:
        sheet = _resolve_sheet(wb, sheet_name)
        ws = wb[sheet]
        # read_only 模式下 ws.max_row 不一定准；直接迭代
        all_rows: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))

    if len(all_rows) < header_row:
        log.warning(
            "Sheet %r 总行数 %d 不足表头行号 %d，返回空列表",
            sheet,
            len(all_rows),
            header_row,
        )
        return []

    raw_header = all_rows[header_row - 1]
    headers: list[str | None] = [
        None if _is_blank(h) else str(h).strip() for h in raw_header
    ]
    valid_indices = [i for i, h in enumerate(headers) if h is not None]
    if not valid_indices:
        log.warning("Sheet %r 表头行 %d 全为空，返回空列表", sheet, header_row)
        return []

    rows_out: list[dict[str, Any]] = []
    for raw_row in all_rows[header_row:]:
        if all(_is_blank(cell) for cell in raw_row):
            continue
        row_dict: dict[str, Any] = {}
        for idx in valid_indices:
            value = raw_row[idx] if idx < len(raw_row) else None
            # 字符串值统一 strip，避免下游做列名/取值匹配时被首尾空格坑
            if isinstance(value, str):
                value = value.strip()
            # 此处 headers[idx] 一定不为 None（valid_indices 已过滤）
            key = headers[idx]
            assert key is not None
            row_dict[key] = value
        rows_out.append(row_dict)

    log.debug(
        "read_rows: %s [%s] header_row=%d → %d 列 × %d 行",
        path.name,
        sheet,
        header_row,
        len(valid_indices),
        len(rows_out),
    )
    return rows_out
