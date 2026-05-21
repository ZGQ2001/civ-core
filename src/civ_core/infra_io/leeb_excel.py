"""里氏硬度原始数据 / 结果 Excel 导入导出（2026-05-20 固化新格式）。

为什么独立成 leeb_excel.py：
  里氏硬度的「每构件 3 行 9 列」固定结构与通用 excel_reader（按表头取列）差异大，
  独立模块语义清晰。

新格式约定（2026-05-20 锁定，详见 docs/civil_kb/formats/leeb_hardness_excel.md）：
    一个 xlsx 文件 = 一个检测项目实例（如「里氏硬度-D号站房.xlsx」）
    一个 sheet     = 一个检测批（sheet 名 = 批名，如「检测批1」）
    sheet 内每构件 3 行：
        | 序号 | 构件位置 | HL1..HL9 | 厚度 |
        | 1  | 2×H钢柱  | 467..463 | 16   |
        |    |          | 471..465 |      |
        |    |          | 477..462 |      |

  • 第 1 行表头固定；数据从第 2 行起
  • 序号 / 构件位置 / 厚度 仅在每构件首行
  • HL 9 列每行都有值
  • 不再有 mid-sheet 子表头（这是旧报检单格式遗留，现在每批独立 sheet）

兼容旧格式：read_leeb_components 单 sheet API 保留 + 跳子表头逻辑保留，
便于读 D 号站房旧报检单（一 sheet 多批）。

测量角度 (angle_degrees) 不在原始数据 Excel 中；由 UI 全局选择传入。

结果文件格式（write_leeb_results_workbook）：
    一个 xlsx 文件 = 一份计算结果
    每检测批生成 2 个 sheet：
      「<批名>-过程数据」：每测区一行 HL_m/HL_t/HL_a/HL_corr/fb_min/fb_max 及构件聚合
      「<批名>-报告插入表」：仿 D 号站房风格，每构件 3 行（带聚合首行），可直接拷贝粘报告
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from civ_core.domain.calc_schema import (
    LeebHardnessBatch,
    LeebHardnessBatchResult,
    LeebHardnessComponentInput,
    LeebHardnessWorkbook,
    LeebHardnessWorkbookResult,
)
from civ_core.utils.exceptions import InfraIOError, InputError
from civ_core.utils.logger import get_logger

log = get_logger(__name__)


# ── 报检单列位置（按 D 号站房格式锁定；其他格式可扩展为参数）────
_COL_SEQ = 1            # A: 序号
_COL_NAME = 2           # B: 构件位置
_COL_HL_START = 3       # C..K: HL1..HL9 (9 列)
_COL_HL_COUNT = 9
_COL_THICKNESS = 12     # L: 厚度
_COL_BATCH = 16         # P: 检测批


def read_leeb_components(
    path: Path,
    sheet_name: str,
    *,
    rows_per_component: int = 3,
    header_rows: int = 1,
    default_angle_degrees: float = -90.0,
) -> list[LeebHardnessComponentInput]:
    """从报检单 Excel 读取多构件里氏硬度数据。

    参数:
        path: xlsx 文件路径
        sheet_name: 工作表名（如「里氏硬度（钢柱）」）
        rows_per_component: 每构件占多少行（= 测区数；规范 ≥3，常用 3）
        header_rows: 表头占多少行（用于偏移）
        default_angle_degrees: 默认角度（钢柱常用 -90°/+90°；UI 全局选）

    返回:
        LeebHardnessComponentInput 列表，按 Excel 序号顺序。

    异常:
        InfraIOError —— 文件不存在 / sheet 不存在 / 文件被占用
        InputError   —— 数据格式不符（HL 非整数 / 9 列缺失 / 厚度非数）
    """
    if not path.exists():
        raise InfraIOError(
            cause=f"文件不存在：{path}",
            location="read_leeb_components",
            hint="请检查路径拼写或文件是否被移动",
        )

    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        raise InfraIOError(
            cause=f"打开 Excel 文件失败：{e}",
            location="read_leeb_components",
            hint="请检查文件是否被 Office 占用、是否为有效 xlsx 格式",
        ) from e

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise InputError(
            cause=f"工作表 {sheet_name!r} 不存在；可选 sheet：{wb.sheetnames}",
            location="read_leeb_components",
            hint="请在文件中确认工作表名称（注意全角/半角括号）",
        )

    ws = wb[sheet_name]
    components: list[LeebHardnessComponentInput] = []
    # 当前检测批名跨子表头继承（每构件首行可能不再写检测批名，但属于上一批）
    current_batch_name = ""

    # 从 header_rows + 1 开始遍历；每 rows_per_component 行处理一个构件
    row = header_rows + 1
    while row <= ws.max_row:
        seq_cell = ws.cell(row, _COL_SEQ).value
        # 跳过空行（报检单底部可能有备注/分隔空行）
        if seq_cell is None and ws.cell(row, _COL_HL_START).value is None:
            row += 1
            continue
        # 跳过 mid-sheet 子表头：报检单每检测批起始处会重复表头行
        # 识别方式：A 列非数字（"序号" 字面量）或 C 列非数字（"里氏硬度值（HLi）"）
        if seq_cell is not None and not isinstance(seq_cell, (int, float)):
            log.debug("跳过子表头 row=%d (A=%r)", row, seq_cell)
            row += 1
            continue
        if not isinstance(ws.cell(row, _COL_HL_START).value, (int, float)):
            log.debug("跳过子表头/异常行 row=%d (C=%r)", row, ws.cell(row, _COL_HL_START).value)
            row += 1
            continue

        name = ws.cell(row, _COL_NAME).value
        thickness = ws.cell(row, _COL_THICKNESS).value
        batch_cell = ws.cell(row, _COL_BATCH).value
        if batch_cell:
            current_batch_name = str(batch_cell).strip()
        batch_name = current_batch_name

        # 读取 rows_per_component 行的 HL 数据
        test_areas: list[tuple[int, ...]] = []
        for offset in range(rows_per_component):
            r = row + offset
            if r > ws.max_row:
                break
            hl_values: list[int] = []
            for c_offset in range(_COL_HL_COUNT):
                v = ws.cell(r, _COL_HL_START + c_offset).value
                if v is None:
                    raise InputError(
                        cause=f"行 {r} 列 {get_column_letter(_COL_HL_START + c_offset)} HL 值缺失",
                        location="read_leeb_components",
                        hint=f"构件 #{seq_cell} 的第 {offset + 1} 测区缺测点；请补完 9 个值",
                    )
                try:
                    hl_values.append(int(round(float(v))))
                except (TypeError, ValueError) as e:
                    raise InputError(
                        cause=f"行 {r} 列 {get_column_letter(_COL_HL_START + c_offset)} HL 值非数字：{v!r}",
                        location="read_leeb_components",
                    ) from e
            test_areas.append(tuple(hl_values))

        if len(test_areas) != rows_per_component:
            log.warning(
                "构件 #%s 实际只读到 %d 个测区（期望 %d），可能文件末尾不完整",
                seq_cell,
                len(test_areas),
                rows_per_component,
            )

        # seq/name/thickness 解析（容忍 None 缺失，用合理默认或抛错）
        try:
            seq = int(seq_cell) if seq_cell is not None else len(components) + 1
        except (TypeError, ValueError) as e:
            raise InputError(
                cause=f"行 {row} 序号非整数：{seq_cell!r}",
                location="read_leeb_components",
            ) from e
        if not name:
            raise InputError(
                cause=f"行 {row} 构件位置（B 列）为空",
                location="read_leeb_components",
                hint=f"序号 {seq} 缺构件位置；请检查 Excel",
            )
        try:
            thickness_f = float(thickness) if thickness is not None else 0.0
        except (TypeError, ValueError) as e:
            raise InputError(
                cause=f"行 {row} 厚度非数字：{thickness!r}",
                location="read_leeb_components",
            ) from e
        if thickness_f <= 0:
            raise InputError(
                cause=f"行 {row} 厚度无效（{thickness_f}），需 > 0",
                location="read_leeb_components",
                hint=f"序号 {seq} 厚度列（L 列）缺失或非正数",
            )

        components.append(
            LeebHardnessComponentInput(
                seq=seq,
                name=str(name).strip(),
                thickness=thickness_f,
                angle_degrees=default_angle_degrees,
                test_areas_raw=tuple(test_areas),
                batch_name=batch_name,
            )
        )

        row += rows_per_component

    wb.close()

    if not components:
        raise InputError(
            cause=f"工作表 {sheet_name!r} 未读到任何构件数据",
            location="read_leeb_components",
            hint=f"请确认数据从第 {header_rows + 1} 行起、每 {rows_per_component} 行一个构件",
        )

    log.info("读取里氏硬度数据完成：%d 个构件（来自 %s / %s）", len(components), path.name, sheet_name)
    return components


# ════════════════════════════════════════════════════════════════
# 新格式：read_leeb_workbook 一文件多批 sheet
# ════════════════════════════════════════════════════════════════
def read_leeb_workbook(
    path: Path,
    *,
    default_angle_degrees: float = 0.0,
    sheet_name_filter: str | None = None,
) -> LeebHardnessWorkbook:
    """读整个 xlsx → LeebHardnessWorkbook（每 sheet = 一个检测批）。

    参数:
        path: xlsx 文件路径
        default_angle_degrees: 全部构件的默认测量角度（UI 后续可整体改）
        sheet_name_filter: 若给则只读名字含该子串的 sheet（例如 "检测批" 跳过元信息 sheet）

    返回:
        LeebHardnessWorkbook：含 N 个 LeebHardnessBatch，按 sheet 顺序保持

    异常:
        InfraIOError —— 文件不存在 / 打不开
        InputError   —— 文件不含任何可解析的 sheet
    """
    if not path.exists():
        raise InfraIOError(
            cause=f"文件不存在：{path}",
            location="read_leeb_workbook",
        )
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
    except Exception as e:
        raise InfraIOError(
            cause=f"打开 Excel 文件失败：{e}",
            location="read_leeb_workbook",
        ) from e

    target_sheets = [
        s for s in sheet_names
        if sheet_name_filter is None or sheet_name_filter in s
    ]
    if not target_sheets:
        raise InputError(
            cause=f"未找到可解析的 sheet（可选：{sheet_names}，过滤：{sheet_name_filter!r}）",
            location="read_leeb_workbook",
            hint="请确认 sheet 命名（如「检测批1」）",
        )

    batches: list[LeebHardnessBatch] = []
    for sheet in target_sheets:
        try:
            components = read_leeb_components(
                path, sheet, default_angle_degrees=default_angle_degrees
            )
        except InputError as e:
            # sheet 完全没数据：跳过（可能是元信息 sheet 比如「委托信息」）
            log.warning("跳过 sheet %r：%s", sheet, e)
            continue
        # 把 sheet 名作为批名注入每个构件（覆盖原 batch_name），便于追溯
        components_with_batch = [
            LeebHardnessComponentInput(
                seq=c.seq,
                name=c.name,
                thickness=c.thickness,
                angle_degrees=c.angle_degrees,
                test_areas_raw=c.test_areas_raw,
                batch_name=sheet,
            )
            for c in components
        ]
        batches.append(LeebHardnessBatch(batch_name=sheet, components=tuple(components_with_batch)))

    if not batches:
        raise InputError(
            cause=f"{path.name} 中所有 sheet 都没有有效数据",
            location="read_leeb_workbook",
            hint="请按格式规范填写至少一个检测批 sheet",
        )

    log.info("读取里氏硬度 workbook 完成：%s → %d 个检测批", path.name, len(batches))
    return LeebHardnessWorkbook(
        batches=tuple(batches),
        file_label=path.stem,
    )


# ════════════════════════════════════════════════════════════════
# 导出：批级结果 → Excel 两张表
# ════════════════════════════════════════════════════════════════
def write_leeb_results(
    path: Path,
    batch: LeebHardnessBatchResult,
    *,
    angle_degrees: float | None = None,
) -> None:
    """把批级结果导出为 xlsx 两张表（原始数据 + 计算结果 + 底部批级汇总）。

    参数:
        path: 输出文件路径（含 .xlsx 后缀）
        batch: calc_leeb_hardness_batch 返回的批级结果
        angle_degrees: 用于在表头追加角度信息（None 则不写）
    """
    wb = Workbook()

    # ── Sheet 1：原始数据（仿报检单结构）───────────────────────
    ws_raw = wb.active
    ws_raw.title = "原始数据"
    ws_raw.append(
        ["序号", "构件位置", *[f"HL{i + 1}" for i in range(9)], "厚度(mm)", "检测批", "角度°"]
    )
    for comp, _result in batch.components_with_results:
        for zone_idx, zone in enumerate(comp.test_areas_raw):
            if zone_idx == 0:
                ws_raw.append(
                    [
                        comp.seq,
                        comp.name,
                        *list(zone),
                        comp.thickness,
                        comp.batch_name,
                        comp.angle_degrees,
                    ]
                )
            else:
                ws_raw.append(["", "", *list(zone), "", "", ""])

    # ── Sheet 2：计算结果（每测区一行 + 构件聚合 + 批级总）──
    ws_res = wb.create_sheet("计算结果")
    ws_res.append(
        [
            "序号",
            "构件位置",
            "测区",
            "HL_m",
            "HL_t",
            "HL_a",
            "HL_corr",
            "fb_min(MPa)",
            "fb_max(MPa)",
            "构件下限平均",
            "构件推定值",
        ]
    )
    for comp, result in batch.components_with_results:
        for zone_idx, area in enumerate(result.test_areas):
            ws_res.append(
                [
                    comp.seq if zone_idx == 0 else "",
                    comp.name if zone_idx == 0 else "",
                    f"测区{zone_idx + 1}",
                    area.hl_m,
                    round(area.hl_t, 2),
                    round(area.hl_a, 2),
                    round(area.hl_corrected, 2),
                    round(area.fb_min, 1),
                    round(area.fb_max, 1),
                    round(result.comp_fb_min_avg, 1) if zone_idx == 0 else "",
                    round(result.comp_fb_est, 1) if zone_idx == 0 else "",
                ]
            )

    # 批级汇总（最后追加 2 行：分隔 + 数据）
    ws_res.append([])
    ws_res.append(
        [
            "—",
            f"批级汇总（共 {batch.n_components} 个构件）",
            "—",
            "—",
            "—",
            "—",
            "—",
            "—",
            "—",
            "批级特征值平均",
            round(batch.batch_fb_char_avg, 1),
        ]
    )

    # 角度元信息追加到表 1 顶部（A 列前可看）
    if angle_degrees is not None:
        ws_raw.cell(1, 14).value = f"全局测量角度：{angle_degrees}°"

    try:
        wb.save(str(path))
    except Exception as e:
        raise InfraIOError(
            cause=f"写入 Excel 失败：{e}",
            location="write_leeb_results",
            hint="请确认目标目录可写且文件未被 Excel 打开",
        ) from e

    log.info(
        "导出里氏硬度结果完成：%s（%d 构件，批级 fb_char_avg=%.1f MPa）",
        path,
        batch.n_components,
        batch.batch_fb_char_avg,
    )


# ════════════════════════════════════════════════════════════════
# 新格式：结果导出（每批 2 sheet：过程 + 报告插入表）
# ════════════════════════════════════════════════════════════════
# 报告插入表样式（仿计算表格.xlsx BH..BO 区域）：
#   检测部位 | 测区 | 测区里氏硬度平均值HL | 修正后里氏硬度平均值HL |
#   抗拉强度最小值fb,min | 抗拉强度特征值N/mm² | 抗拉强度特征值平均值N/mm²
# 每构件 3 行（3 测区），首行带构件名 + 构件级聚合值
_REPORT_HEADER = (
    "检测部位",
    "测区",
    "测区里氏硬度平均值 (HL)",
    "修正后里氏硬度平均值 (HL)",
    "抗拉强度最小值 fb,min (N/mm²)",
    "抗拉强度特征值 (N/mm²)",
    "抗拉强度特征值平均值 (N/mm²)",
)

_PROCESS_HEADER = (
    "序号",
    "构件位置",
    "测区",
    "HL_m",
    "HL_t (厚度修正)",
    "HL_a (角度修正)",
    "HL_corr",
    "fb_min (MPa)",
    "fb_max (MPa)",
    "构件下限平均",
    "构件上限平均",
    "构件推定值",
)


def _bold_header(ws, row: int = 1) -> None:
    """给指定行所有非空单元格加粗 + 浅灰底色 + 居中。"""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="E8EEF5")
    align = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = bold
            cell.fill = fill
            cell.alignment = align


def _write_process_sheet(ws, batch_result: LeebHardnessBatchResult, angle_degrees: float) -> None:
    """写「<批名>-过程数据」sheet：每测区一行，含全部中间量 + 构件聚合 + 底部批级。"""
    ws.append(list(_PROCESS_HEADER))
    for comp, result in batch_result.components_with_results:
        for zone_idx, area in enumerate(result.test_areas):
            ws.append(
                [
                    comp.seq if zone_idx == 0 else "",
                    comp.name if zone_idx == 0 else "",
                    f"测区{zone_idx + 1}",
                    area.hl_m,
                    round(area.hl_t, 2),
                    round(area.hl_a, 2),
                    round(area.hl_corrected, 2),
                    round(area.fb_min, 1),
                    round(area.fb_max, 1),
                    round(result.comp_fb_min_avg, 1) if zone_idx == 0 else "",
                    round(result.comp_fb_max_avg, 1) if zone_idx == 0 else "",
                    round(result.comp_fb_est, 1) if zone_idx == 0 else "",
                ]
            )
    # 底部批级汇总
    ws.append([])
    ws.append(
        [
            "—", f"批级汇总（{batch_result.n_components} 构件，角度 {angle_degrees}°）",
            "—", "—", "—", "—", "—", "—", "—",
            "批级特征值平均",
            "—",
            round(batch_result.batch_fb_char_avg, 1),
        ]
    )
    _bold_header(ws)


def _write_report_sheet(ws, batch_result: LeebHardnessBatchResult) -> None:
    """写「<批名>-报告插入表」sheet：仿 D 号站房报告风格，可直接拷贝粘报告。"""
    ws.append(list(_REPORT_HEADER))
    for _comp, result in batch_result.components_with_results:
        # 首行带构件名 + 构件级聚合
        for zone_idx, area in enumerate(result.test_areas):
            row = [
                _comp.name if zone_idx == 0 else "",
                f"测区{zone_idx + 1}",
                area.hl_m,
                round(area.hl_corrected, 1),
                round(area.fb_min, 1),
                round(result.comp_fb_min_avg, 1) if zone_idx == 0 else "",
                # 抗拉强度特征值平均值：所有 fb_min 单值再做构件平均（同 comp_fb_min_avg）
                round(result.comp_fb_min_avg, 1) if zone_idx == 0 else "",
            ]
            ws.append(row)
    # 底部追加批级特征值平均行（合并第 7 列右侧显示）
    ws.append([])
    ws.append([
        "批级特征值平均", "", "", "", "", "",
        round(batch_result.batch_fb_char_avg, 1),
    ])
    _bold_header(ws)


def write_leeb_results_workbook(
    path: Path,
    workbook_result: LeebHardnessWorkbookResult,
    *,
    angle_degrees: float,
    include_report_sheet: bool = True,
) -> None:
    """把 workbook 结果导出为 xlsx：每批 1 或 2 sheet。

    include_report_sheet=True（默认，CLI / 旧调用兼容）：每批 2 sheet（过程 + 报告插入表）
    include_report_sheet=False（RPC handler 用）：每批只写「过程数据」sheet；
      「报告插入表」交给 C# sidecar (xlsx.write_leeb_report_table) 用 ClosedXML 生成
      精致格式（合并单元格 / 字体 / 边框 / 列宽）追加到同文件。

    sheet 命名：「<批名>-过程数据」+ 可选「<批名>-报告插入表」
    sheet 顺序：按批分组（A 批过程、A 批报告、B 批过程、B 批报告）
    """
    wb = Workbook()
    # 删除默认空 sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    for br in workbook_result.batch_results:
        # sheet 名长度上限 31 字符（openpyxl/Excel 硬限）
        proc_name = _safe_sheet_name(f"{br.batch_name}-过程数据")
        ws_proc = wb.create_sheet(proc_name)
        _write_process_sheet(ws_proc, br, angle_degrees)
        if include_report_sheet:
            rep_name = _safe_sheet_name(f"{br.batch_name}-报告插入表")
            ws_rep = wb.create_sheet(rep_name)
            _write_report_sheet(ws_rep, br)

    try:
        wb.save(str(path))
    except Exception as e:
        raise InfraIOError(
            cause=f"写入 Excel 失败：{e}",
            location="write_leeb_results_workbook",
            hint="请确认目标目录可写且文件未被 Excel 打开",
        ) from e

    log.info(
        "导出里氏硬度 workbook 结果：%s（%d 批，共 %d 构件）",
        path,
        workbook_result.n_batches,
        workbook_result.n_components_total,
    )


def _safe_sheet_name(name: str) -> str:
    """Excel sheet 名限制：≤31 字符，不能含 / \\ ? * [ ] :。"""
    bad_chars = '/\\?*[]:'
    cleaned = "".join("_" if c in bad_chars else c for c in name)
    return cleaned[:31]
