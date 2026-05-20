"""里氏硬度（INSP-001）批级计算工具页。

工作流（钢结构厂房项目实战）：
  1. [导入 Excel] 选「里氏硬度（钢柱/钢梁）」sheet → 解析为构件清单
  2. 顶栏选全局测量角度（-90/-45/0/+45/+90）
  3. [计算] → 调 calc_leeb_hardness_batch → 右栏显示结果
  4. [导出 Excel] → 输出原始数据 + 计算结果两张 sheet 用于贴报告

布局：
  ┌────────── 顶栏（导入 / 角度 / 计算 / 导出 / 清空） ──────────┐
  ├──────────┬──────────────────────────────────────────────────┤
  │ 构件清单  │ 批级 fb_char_avg（醒目大字号）                    │
  │ Table    │ ────────────────────────────────                  │
  │ (40%)    │ 详细结果 Table（每测区一行）                       │
  │          │ (60%)                                            │
  └──────────┴──────────────────────────────────────────────────┘

设计要点：
  • 数据持有：self._components / self._batch_result，UI 模型从这两个状态读
  • 不依赖 worker thread：批级计算 28 构件 × 3 测区 ≈ 60ms，主线程同步即可
  • angle_degrees 是全局参数（钢结构常用 90°/向下垂直），UI 顶栏下拉一次性设
  • db 由外部（MainWindow）注入，避免本视图自己 init standards.db
"""

from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import QAbstractTableModel, QModelIndex, Qt
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QSplitter,
    QTableView,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import (
    BodyLabel,
    ComboBox,
    PrimaryPushButton,
    PushButton,
    StrongBodyLabel,
    TitleLabel,
)

from civ_core.core.calc_functions import calc_leeb_hardness_batch
from civ_core.domain.calc_schema import (
    LeebHardnessBatchResult,
    LeebHardnessComponentInput,
)
from civ_core.infra_io.leeb_excel import (
    read_leeb_components,
    write_leeb_results,
)
from civ_core.infra_io.standards_db import StandardsDB
from civ_core.ui.components.error_infobar import (
    show_error_infobar,
    show_success_infobar,
    show_warning_infobar,
)
from civ_core.utils.exceptions import CivCoreError
from civ_core.utils.logger import get_logger

log = get_logger(__name__)


# ── 表格模型 ────────────────────────────────────────────────────
class _ComponentsTableModel(QAbstractTableModel):
    """左栏构件清单：序号 / 构件位置 / 厚度 / 测区数 / 检测批。"""

    HEADERS = ("序号", "构件位置", "厚度(mm)", "测区数", "检测批")

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._data: list[LeebHardnessComponentInput] = []

    def set_components(self, components: list[LeebHardnessComponentInput]) -> None:
        self.beginResetModel()
        self._data = components
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: B008
        return len(self._data)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: B008
        return len(self.HEADERS)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> object:
        if not index.isValid() or role != Qt.ItemDataRole.DisplayRole:
            return None
        c = self._data[index.row()]
        col = index.column()
        if col == 0:
            return c.seq
        if col == 1:
            return c.name
        if col == 2:
            return f"{c.thickness:g}"
        if col == 3:
            return len(c.test_areas_raw)
        if col == 4:
            return c.batch_name
        return None

    def headerData(
        self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole
    ) -> object:
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.HEADERS[section]
        return None


class _ResultsTableModel(QAbstractTableModel):
    """右栏详细结果：每测区一行 + 构件聚合在首行显示。"""

    HEADERS = (
        "序号",
        "构件位置",
        "测区",
        "HL_m",
        "HL_t",
        "HL_a",
        "HL_corr",
        "fb_min",
        "fb_max",
        "构件推定",
    )

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        # 行：tuple (comp_seq | "", comp_name | "", zone_label, hl_m, hl_t, hl_a, hl_corr, fb_min, fb_max, comp_est | "")
        self._rows: list[tuple] = []

    def set_batch(self, batch: LeebHardnessBatchResult | None) -> None:
        self.beginResetModel()
        self._rows = []
        if batch is not None:
            for comp, result in batch.components_with_results:
                for zone_idx, area in enumerate(result.test_areas):
                    self._rows.append(
                        (
                            comp.seq if zone_idx == 0 else "",
                            comp.name if zone_idx == 0 else "",
                            f"测区{zone_idx + 1}",
                            area.hl_m,
                            f"{area.hl_t:.2f}",
                            f"{area.hl_a:.2f}",
                            f"{area.hl_corrected:.2f}",
                            f"{area.fb_min:.1f}",
                            f"{area.fb_max:.1f}",
                            f"{result.comp_fb_est:.1f}" if zone_idx == 0 else "",
                        )
                    )
        self.endResetModel()

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: B008
        return len(self._rows)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: B008
        return len(self.HEADERS)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole) -> object:
        if not index.isValid() or role != Qt.ItemDataRole.DisplayRole:
            return None
        return self._rows[index.row()][index.column()]

    def headerData(
        self, section: int, orientation: Qt.Orientation, role: int = Qt.ItemDataRole.DisplayRole
    ) -> object:
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self.HEADERS[section]
        return None


# ── 主视图 ──────────────────────────────────────────────────────
# 角度档：UI 用整数度数（与规范表 key1 一致）
_ANGLE_OPTIONS: tuple[tuple[str, float], ...] = (
    ("-90° 向上垂直", -90.0),
    ("-45° 向上 45°", -45.0),
    ("0° 水平", 0.0),
    ("+45° 向下 45°", 45.0),
    ("+90° 向下垂直", 90.0),
)


class LeebHardnessView(QWidget):
    """里氏硬度批级计算工具页（导航 routing key = leebHardnessPage）。"""

    def __init__(self, db: StandardsDB, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("leebHardnessPage")
        self._db = db
        self._components: list[LeebHardnessComponentInput] = []
        self._batch_result: LeebHardnessBatchResult | None = None
        self._last_import_dir: Path | None = None

        self._build_ui()
        self._refresh_buttons()

    # ── UI 构造 ──────────────────────────────────────────────
    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(10)

        # ── 顶栏：操作按钮 ───────────────────────────────────
        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        self.btn_import = PushButton("导入 Excel")
        self.btn_import.clicked.connect(self._on_import)
        toolbar.addWidget(self.btn_import)

        toolbar.addWidget(BodyLabel("测量角度："))
        self.cmb_angle = ComboBox()
        for label, _ in _ANGLE_OPTIONS:
            self.cmb_angle.addItem(label)
        # 钢结构常用 +90°（向下垂直）—— 找索引 4
        self.cmb_angle.setCurrentIndex(4)
        self.cmb_angle.currentIndexChanged.connect(self._on_angle_changed)
        toolbar.addWidget(self.cmb_angle)

        self.btn_calc = PrimaryPushButton("▶ 计算")
        self.btn_calc.clicked.connect(self._on_calculate)
        toolbar.addWidget(self.btn_calc)

        self.btn_export = PushButton("导出 Excel")
        self.btn_export.clicked.connect(self._on_export)
        toolbar.addWidget(self.btn_export)

        self.btn_clear = PushButton("清空")
        self.btn_clear.clicked.connect(self._on_clear)
        toolbar.addWidget(self.btn_clear)

        toolbar.addStretch(1)
        self.lbl_status = BodyLabel("尚未导入数据")
        toolbar.addWidget(self.lbl_status)

        root.addLayout(toolbar)

        # ── 主体：左右两栏 ───────────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal, self)
        splitter.setHandleWidth(4)

        # 左：构件清单
        left = QWidget()
        left_lo = QVBoxLayout(left)
        left_lo.setContentsMargins(0, 0, 0, 0)
        left_lo.addWidget(StrongBodyLabel("构件清单"))
        self.tbl_components = QTableView()
        self._comp_model = _ComponentsTableModel(self)
        self.tbl_components.setModel(self._comp_model)
        self.tbl_components.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.tbl_components.horizontalHeader().setStretchLastSection(True)
        self.tbl_components.verticalHeader().setVisible(False)
        self.tbl_components.setAlternatingRowColors(True)
        left_lo.addWidget(self.tbl_components, stretch=1)
        splitter.addWidget(left)

        # 右：批级摘要 + 详细结果
        right = QWidget()
        right_lo = QVBoxLayout(right)
        right_lo.setContentsMargins(0, 0, 0, 0)

        # 批级 fb_char_avg 醒目显示
        self.lbl_batch_summary = TitleLabel("批级抗拉强度特征值平均：—")
        self.lbl_batch_summary.setFont(self._make_summary_font())
        right_lo.addWidget(self.lbl_batch_summary)

        right_lo.addWidget(StrongBodyLabel("详细结果（每测区一行）"))
        self.tbl_results = QTableView()
        self._res_model = _ResultsTableModel(self)
        self.tbl_results.setModel(self._res_model)
        self.tbl_results.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self.tbl_results.verticalHeader().setVisible(False)
        self.tbl_results.setAlternatingRowColors(True)
        right_lo.addWidget(self.tbl_results, stretch=1)

        splitter.addWidget(right)
        splitter.setSizes([400, 600])
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)
        root.addWidget(splitter, stretch=1)

    def _make_summary_font(self) -> QFont:
        f = QFont()
        f.setPointSize(16)
        f.setBold(True)
        return f

    # ── 状态切换 ─────────────────────────────────────────────
    def _refresh_buttons(self) -> None:
        has_data = bool(self._components)
        has_result = self._batch_result is not None
        self.btn_calc.setEnabled(has_data)
        self.btn_export.setEnabled(has_result)
        self.btn_clear.setEnabled(has_data or has_result)

    # ── 事件处理 ─────────────────────────────────────────────
    def _on_import(self) -> None:
        """选择 Excel 文件 → 选 sheet → 解析为构件清单。"""
        start_dir = str(self._last_import_dir) if self._last_import_dir else ""
        path_str, _ = QFileDialog.getOpenFileName(
            self,
            "选择里氏硬度报检单",
            start_dir,
            "Excel 文件 (*.xlsx *.xlsm);;所有文件 (*.*)",
        )
        if not path_str:
            return
        path = Path(path_str)
        self._last_import_dir = path.parent

        # 预读 sheet 列表让用户选
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True)
            sheet_names = list(wb.sheetnames)
            wb.close()
        except Exception as e:
            show_error_infobar(
                self,
                CivCoreError(cause=f"读取 sheet 列表失败：{e}", location="import"),
                where="导入",
            )
            return

        if not sheet_names:
            show_warning_infobar(self, "无工作表", "Excel 文件不含任何工作表")
            return

        # 优先列出含「里氏硬度」的 sheet
        leeb_sheets = [s for s in sheet_names if "里氏" in s or "硬度" in s]
        candidates = leeb_sheets + [s for s in sheet_names if s not in leeb_sheets]

        sheet_name, ok = QInputDialog.getItem(
            self,
            "选择工作表",
            f"共 {len(candidates)} 张工作表，请选择里氏硬度数据所在的表：",
            candidates,
            0,
            False,
        )
        if not ok or not sheet_name:
            return

        # 解析
        try:
            angle = self._current_angle()
            components = read_leeb_components(
                path, sheet_name, default_angle_degrees=angle
            )
        except CivCoreError as e:
            show_error_infobar(self, e, where="导入")
            return
        except Exception as e:
            show_error_infobar(
                self,
                CivCoreError(cause=f"解析失败：{e}", location="read_leeb_components"),
                where="导入",
            )
            log.exception("导入 Excel 失败")
            return

        self._components = components
        self._batch_result = None
        self._comp_model.set_components(components)
        self._res_model.set_batch(None)
        self.lbl_batch_summary.setText("批级抗拉强度特征值平均：— （请点 ▶ 计算）")
        self.lbl_status.setText(f"已导入 {len(components)} 个构件（{sheet_name}）")
        self._refresh_buttons()
        show_success_infobar(
            self,
            "导入成功",
            f"{path.name} / {sheet_name} → {len(components)} 个构件",
        )

    def _on_angle_changed(self) -> None:
        """角度切换 → 同步更新已导入构件的 angle_degrees。"""
        if not self._components:
            return
        angle = self._current_angle()
        new_components: list[LeebHardnessComponentInput] = []
        for c in self._components:
            new_components.append(
                LeebHardnessComponentInput(
                    seq=c.seq,
                    name=c.name,
                    thickness=c.thickness,
                    angle_degrees=angle,
                    test_areas_raw=c.test_areas_raw,
                    batch_name=c.batch_name,
                )
            )
        self._components = new_components
        # 结果作废，需要重算
        self._batch_result = None
        self._res_model.set_batch(None)
        self.lbl_batch_summary.setText("批级抗拉强度特征值平均：— （角度已变，请重新计算）")
        self._refresh_buttons()

    def _current_angle(self) -> float:
        return _ANGLE_OPTIONS[self.cmb_angle.currentIndex()][1]

    def _on_calculate(self) -> None:
        if not self._components:
            return
        try:
            batch = calc_leeb_hardness_batch(self._components, db=self._db)
        except CivCoreError as e:
            show_error_infobar(self, e, where="计算")
            return
        except Exception as e:
            show_error_infobar(
                self,
                CivCoreError(cause=f"计算失败：{e}", location="calc_leeb_hardness_batch"),
                where="计算",
            )
            log.exception("批级计算失败")
            return

        self._batch_result = batch
        self._res_model.set_batch(batch)
        self.lbl_batch_summary.setText(
            f"批级抗拉强度特征值平均：{batch.batch_fb_char_avg:.1f} MPa  "
            f"（共 {batch.n_components} 个构件）"
        )
        self.lbl_status.setText(
            f"已计算 {batch.n_components} 个构件 / {sum(len(c.test_areas_raw) for c in self._components)} 测区"
        )
        self._refresh_buttons()

    def _on_export(self) -> None:
        if self._batch_result is None:
            return
        start_dir = str(self._last_import_dir) if self._last_import_dir else ""
        path_str, _ = QFileDialog.getSaveFileName(
            self,
            "导出里氏硬度计算结果",
            start_dir + "/里氏硬度_计算结果.xlsx" if start_dir else "里氏硬度_计算结果.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not path_str:
            return
        path = Path(path_str)
        try:
            write_leeb_results(path, self._batch_result, angle_degrees=self._current_angle())
        except CivCoreError as e:
            show_error_infobar(self, e, where="导出")
            return

        show_success_infobar(self, "导出成功", f"已写入 {path.name}")

    def _on_clear(self) -> None:
        self._components = []
        self._batch_result = None
        self._comp_model.set_components([])
        self._res_model.set_batch(None)
        self.lbl_batch_summary.setText("批级抗拉强度特征值平均：—")
        self.lbl_status.setText("尚未导入数据")
        self._refresh_buttons()
