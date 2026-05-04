"""Excel（pandas / openpyxl）通用 IO（无 UI 依赖）。

工程规范落地：
  ✓ logger（不再 print）
  ✓ openpyxl 用 with 关闭 workbook
  ✓ 返回 ExcelReplaceResult dataclass
  ✓ 异常带上下文，抛 IOReadError 让 UI 友好提示
"""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from civil_auto.models.schema import ExcelReplaceResult, IOReadError
from civil_auto.utils.logger import get_logger
from civil_auto.utils.patterns import FIG_PATTERN
from civil_auto.utils.word_helpers import make_caption_substitutor

log = get_logger(__name__)


# ──────────────────────────────────────────────────────────────────
# 1. 元信息读取
# ──────────────────────────────────────────────────────────────────
def get_excel_sort_order(
    excel_path: Path | str,
    col_name: str,
    sheet_name: str | None = None,
) -> list[int]:
    """读 Excel 指定列，把每行的「图 N」提成数字列表，作为排序顺序。"""
    path = Path(excel_path)
    if not path.is_file():
        raise IOReadError(f"Excel 文件不存在: {path}", hint=f"请检查路径：{path}")

    try:
        import pandas as pd
    except ImportError as e:
        raise IOReadError(f"缺失 pandas: {e}") from e

    try:
        df = pd.read_excel(path, sheet_name=sheet_name) if sheet_name else pd.read_excel(path)
    except Exception as e:
        raise IOReadError(
            f"读取 Excel 失败 ({path.name}, sheet={sheet_name}): {e}",
            hint="请确认 sheet 名拼写正确，且文件未被其他程序锁定。",
        ) from e

    if col_name not in df.columns:
        raise IOReadError(
            f"Sheet [{sheet_name or '默认'}] 中找不到列：{col_name}",
            hint=f"实际可用列：{list(df.columns)}",
        )

    order: list[int] = []
    for item in df[col_name].dropna().astype(str).tolist():
        m = FIG_PATTERN.search(item)
        if m:
            order.append(int(m.group(1)))
    log.debug("从 [%s].%s 抽取到 %d 个排序键", path.name, col_name, len(order))
    return order


def find_column_index(ws: Worksheet, col_name: str, header_row: int = 1) -> int | None:
    """在指定 sheet 的表头行里找列名对应的列号（1-indexed），找不到返回 None。"""
    for cell in ws[header_row]:
        if cell.value is not None and str(cell.value).strip() == col_name:
            return cell.column
    return None


# ──────────────────────────────────────────────────────────────────
# 2. workbook 资源管理
# ──────────────────────────────────────────────────────────────────
@contextmanager
def open_workbook(path: Path | str, *, read_only: bool = False) -> Iterator[Workbook]:
    """openpyxl Workbook 的 with 包装：确保异常时也 close。

    openpyxl 的 Workbook 持有文件句柄；忘记 close 在 Windows 下会导致
    后续 SaveAs 报「文件被占用」。
    """
    p = Path(path)
    if not p.is_file():
        raise IOReadError(f"Excel 文件不存在: {p}", hint=f"路径：{p}")

    try:
        wb = load_workbook(str(p), read_only=read_only)
    except Exception as e:
        raise IOReadError(f"无法打开 Excel ({p.name}): {e}") from e

    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception as e:
            log.warning("关闭 workbook 时出错（已忽略，文件可能已 save）: %s", e)


# ──────────────────────────────────────────────────────────────────
# 3. 列内容批量替换
# ──────────────────────────────────────────────────────────────────
def replace_in_excel_column(
    excel_path: Path | str,
    sheet_name: str | None,
    col_name: str,
    mapping: dict[int, int],
    output_path: Path | str,
    header_row: int = 1,
) -> ExcelReplaceResult:
    """用 openpyxl 直改指定列的「图 N」，保留所有其他格式。"""
    src = Path(excel_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with open_workbook(src) as wb:
        target_sheet = (
            sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        )
        ws = wb[target_sheet]

        col_idx = find_column_index(ws, col_name, header_row)
        if col_idx is None:
            raise IOReadError(
                f"Sheet [{target_sheet}] 中找不到列：{col_name}",
                hint=f"请确认列名拼写。前几列：{[c.value for c in ws[header_row][:6]]}",
            )

        apply, unmatched = make_caption_substitutor(mapping)

        replaced = 0
        for row in ws.iter_rows(min_row=header_row + 1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value is None:
                continue
            s = str(cell.value)
            if not FIG_PATTERN.search(s):
                continue
            new_s = apply(s)
            if new_s != s:
                cell.value = new_s
                replaced += 1

        wb.save(str(out))

    log.info(
        "Excel 列替换完成: sheet=%s, col=%s, 替换 %d 单元格, 未匹配 %d → %s",
        target_sheet,
        col_name,
        replaced,
        len(unmatched),
        out.name,
    )
    return ExcelReplaceResult(
        output_path=out,
        cells_replaced=replaced,
        unmatched_old_ids=unmatched,
    )
