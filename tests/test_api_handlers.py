"""api handlers (plot_curves) + 端到端 dispatch 测试。

workspace.* / files.* / leeb.* / pdf_tools.* / word2pdf.* 已迁 C# sidecar (civ-doc)
—— 它们的用例在 dotnet/civ-doc.Tests/ 下。Python 端只剩 plot_curves（matplotlib 无可替代）。
"""

from __future__ import annotations

import json

import pytest

from civ_core.api import handlers
from civ_core.api.handlers import plot_curves as plot_handler


# ── 端到端 dispatch + handler 注册 ───────────────────────
def test_full_dispatcher_methods() -> None:
    """build_dispatcher 注册了 plot_curves 全部方法 + ping/version。"""
    from civ_core.api.__main__ import build_dispatcher

    d = build_dispatcher()
    methods = d.methods()
    # 关键方法都在
    for m in (
        "ping",
        "version",
        "plot_curves.list_presets",
        "plot_curves.run",
    ):
        assert m in methods


def test_dispatcher_only_exposes_whitelisted_methods() -> None:
    """register_module 必须只暴露 __all__ 里的方法 —— 避免顶部 import 的
    工具类（Path / dataclass / 业务函数）被误注册成 RPC 方法，造成边界泄漏。"""
    from civ_core.api.__main__ import build_dispatcher

    d = build_dispatcher()
    methods = set(d.methods())
    # 这些是模块顶部 import 进来的，绝不能被注册成 RPC
    forbidden = {
        "plot_curves.Path",
        "plot_curves.PlotCurvesError",
        "plot_curves.get_preset_names",
        "plot_curves.load_presets",
        "plot_curves.run_plot_curves",
    }
    leaked = forbidden & methods
    assert not leaked, f"非业务方法被泄漏为 RPC: {leaked}"


def test_ping_roundtrip_via_dispatcher() -> None:
    from civ_core.api.__main__ import build_dispatcher

    d = build_dispatcher()
    req = json.dumps({"jsonrpc": "2.0", "id": 1, "method": "ping"})
    resp = json.loads(d.handle_raw(req))
    assert resp["result"] == "pong"


def test_handlers_module_exposes_submodules() -> None:
    assert hasattr(handlers, "plot_curves")


# ── plot_curves handler ───────────────────────────────────
def test_plot_curves_list_presets_shape() -> None:
    """list_presets 返回 {presets:[str], default:str|None}；预设库应不为空（系统预设至少 1 条）。"""
    res = plot_handler.list_presets()
    assert isinstance(res["presets"], list)
    assert all(isinstance(n, str) for n in res["presets"])
    # 系统至少有一条预设（healthcheck 验过）
    assert len(res["presets"]) >= 1
    assert res["default"] == res["presets"][0]


def test_plot_curves_run_rejects_missing_preset(tmp_path) -> None:
    """跑不存在的预设 → PlotCurvesError（dispatcher 会包成 -32603）。"""
    from civ_core.core.plot_curves import PlotCurvesError

    fake_xlsx = tmp_path / "x.xlsx"
    fake_xlsx.write_bytes(b"not a real xlsx")  # 不会真的去读到这一步
    with pytest.raises(PlotCurvesError):
        plot_handler.run(
            excel_path=str(fake_xlsx),
            preset="不存在的预设名",
            output_dir=str(tmp_path / "out"),
        )


def test_plot_curves_render_preview_missing_excel(tmp_path) -> None:
    """render_preview 缺 excel → 抛错（不静默返回空字节）。"""
    bad_preset = {
        "id_column": "X",
        "filename_template": "{id}.png",
        "title_template": "{id}",
        "x_axis": {"label": "x", "range": None},
        "y_axis": {"label": "y", "range": None},
        "curves": [],
    }
    with pytest.raises(Exception):
        plot_handler.render_preview(bad_preset, str(tmp_path / "nope.xlsx"))


def test_plot_curves_render_preview_returns_row_data(tmp_path) -> None:
    """render_preview 必须返回当前 row 的所有列值（之前用 stem 反查 id 永远空）。"""
    from openpyxl import Workbook

    xlsx = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.append(["编号", "X值", "Y值", "备注"])
    ws.append(["A1", 1.0, 10.0, "first"])
    ws.append(["A2", 2.0, 20.0, "second"])
    wb.save(str(xlsx))

    # filename_template 含中文后缀 —— stem 不等于 id，原 bug 就是这里失配
    preset = {
        "id_column": "编号",
        "filename_template": "{id}_曲线.png",
        "title_template": "{id} 预览",
        "x_axis": {"label": "X", "range": None},
        "y_axis": {"label": "Y", "range": None},
        "curves": [
            {
                "name": "曲线",
                "color": "#1F4FE0",
                "marker": "o",
                "linewidth": 2.0,
                "markersize": 6,
                "points": [
                    {"fixed_axis": "x", "fixed_value": 1.0, "var_column": "Y值"},
                ],
            }
        ],
    }
    # 第 0 张（A1）的 row_data 应该非空且含原始列
    res0 = plot_handler.render_preview(preset, str(xlsx), row_index=0)
    assert res0["row_data"], "row_data 不能为空 dict"
    assert res0["row_data"]["编号"] == "A1"
    assert res0["row_data"]["X值"] == 1.0
    assert res0["row_data"]["备注"] == "first"

    # 第 1 张（A2）应该是另一行
    res1 = plot_handler.render_preview(preset, str(xlsx), row_index=1)
    assert res1["row_data"]["编号"] == "A2"
    assert res1["row_data"]["备注"] == "second"


def test_plot_curves_render_preview_returns_base64_png(tmp_path) -> None:
    """端到端：写一个最小 xlsx + 最小预设 → 拿到非空 base64 PNG。"""
    import base64

    from openpyxl import Workbook

    xlsx = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.append(["编号", "X值", "Y值"])
    ws.append(["A1", 1.0, 10.0])
    ws.append(["A2", 2.0, 20.0])
    wb.save(str(xlsx))

    preset = {
        "id_column": "编号",
        "filename_template": "{id}.png",
        "title_template": "{id} 预览",
        "x_axis": {"label": "X", "range": None},
        "y_axis": {"label": "Y", "range": None},
        "curves": [
            {
                "name": "曲线",
                "color": "#1F4FE0",
                "marker": "o",
                "linewidth": 2.0,
                "markersize": 6,
                "points": [
                    {"fixed_axis": "x", "fixed_value": 1.0, "var_column": "Y值"},
                ],
            }
        ],
    }
    res = plot_handler.render_preview(preset, str(xlsx))
    assert res["mime"] == "image/png"
    assert res["total_rows"] == 2
    # base64 解出来必须是有效 PNG（前 8 字节为 PNG 魔数）
    raw = base64.b64decode(res["png_base64"])
    assert raw[:8] == b"\x89PNG\r\n\x1a\n"


def test_plot_curves_run_default_output_dir(tmp_path, monkeypatch) -> None:
    """output_dir=None 时默认 <excel 同级>/曲线图/。通过 mock run_plot_curves 验路径计算。"""
    from civ_core.api.handlers import plot_curves as ph
    from civ_core.core.plot_curves import RunResult

    excel = tmp_path / "data.xlsx"
    excel.write_bytes(b"")
    captured = {}

    def fake_run(**kwargs):
        captured.update(kwargs)
        # 返回空结果即可
        from civ_core.core.plot_curves import BuildSummary

        return RunResult(
            written=[],
            failed=[],
            summary=BuildSummary(total_rows=0, skipped_empty_id=[], skipped_bad_data=[]),
        )

    monkeypatch.setattr(ph, "run_plot_curves", fake_run)
    ph.run(excel_path=str(excel), preset="x")  # output_dir 省略
    assert captured["output_dir"] == excel.parent / "曲线图"


# ── leeb handler ──────────────────────────────────────────
# leeb.* 已迁 C# sidecar (civ-doc) —— Python 端不再持有；测试在
# dotnet/civ-doc.Tests/Leeb*Tests.cs（41 个用例对照 Python 黄金值）。


# pdf_tools.* 已迁 C# sidecar (civ-doc) —— 用例在
# dotnet/civ-doc.Tests/PdfToolsHandlersTests.cs。


# word2pdf.* 已迁 C# sidecar (civ-doc) —— 用例在
# dotnet/civ-doc.Tests/Word2PdfHandlersTests.cs。
