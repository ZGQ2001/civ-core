"""ExcelDataCache（L-2.1）单元测试。

为什么需要缓存：
  • LivePreviewPane 防抖后频繁请求重绘，每次都重读 Excel 太浪费（仪器导出
    几万行也常见）
  • 切预设时只重读"映射规则"，不重读 Excel —— 缓存按 (path, mtime,
    sheet_name, header_row) 命中

不测的内容：
  • read_rows 的具体清洗策略（已在 test_basic / infra_io 层测过）
  • Windows 文件锁 / 并发读 —— 单线程缓存，不涉及锁
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import Workbook


def _make_xlsx(path: Path, rows: list[list[object]]) -> None:
    """造一个简易 xlsx 给测试用。rows[0] 当表头。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for r in rows:
        ws.append(r)
    wb.save(str(path))


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "data.xlsx"
    _make_xlsx(
        p,
        [
            ["编号", "位移", "荷载"],
            ["A-01", 0.0, 0.0],
            ["A-02", 1.5, 30.0],
            ["A-03", 3.0, 60.0],
        ],
    )
    return p


# ──────────────────────────────────────────────────────────────────
# 基础：能读、返回的 row 结构正确
# ──────────────────────────────────────────────────────────────────
class TestBasicRead:
    def test_returns_list_of_dicts(self, sample_xlsx: Path) -> None:
        from civ_core.core.data_cache import ExcelDataCache

        cache = ExcelDataCache()
        rows = cache.get_rows(sample_xlsx, sheet_name=None, header_row=1)
        assert isinstance(rows, list)
        assert len(rows) == 3
        assert rows[0] == {"编号": "A-01", "位移": 0.0, "荷载": 0.0}


# ──────────────────────────────────────────────────────────────────
# 缓存命中：同 (path, mtime, sheet, header_row) → 同一对象
# ──────────────────────────────────────────────────────────────────
class TestCacheHit:
    def test_same_key_returns_cached_object(self, sample_xlsx: Path) -> None:
        """命中缓存时，第二次返回的应是同一个 list 对象（is，不只是 ==）。"""
        from civ_core.core.data_cache import ExcelDataCache

        cache = ExcelDataCache()
        r1 = cache.get_rows(sample_xlsx, None, 1)
        r2 = cache.get_rows(sample_xlsx, None, 1)
        assert r1 is r2

    def test_different_sheet_misses_cache(self, tmp_path: Path) -> None:
        """同文件不同 sheet → 不同 key → 各自缓存。"""
        from civ_core.core.data_cache import ExcelDataCache

        p = tmp_path / "twosheets.xlsx"
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "S1"
        ws1.append(["a"])
        ws1.append([1])
        ws2 = wb.create_sheet("S2")
        ws2.append(["b"])
        ws2.append([2])
        wb.save(str(p))

        cache = ExcelDataCache()
        r1 = cache.get_rows(p, sheet_name="S1", header_row=1)
        r2 = cache.get_rows(p, sheet_name="S2", header_row=1)
        assert r1 != r2
        assert r1[0] == {"a": 1}
        assert r2[0] == {"b": 2}

    def test_different_header_row_misses_cache(self, tmp_path: Path) -> None:
        from civ_core.core.data_cache import ExcelDataCache

        p = tmp_path / "hdr.xlsx"
        _make_xlsx(
            p,
            [
                ["meta", "info"],
                ["编号", "位移"],
                ["A-1", 1.0],
            ],
        )
        cache = ExcelDataCache()
        r1 = cache.get_rows(p, None, header_row=1)
        r2 = cache.get_rows(p, None, header_row=2)
        # header_row=1 把第一行 ["meta","info"] 当表头；r2 把第二行当表头
        assert "meta" in r1[0]
        assert "编号" in r2[0]


# ──────────────────────────────────────────────────────────────────
# mtime 变更：触发重读
# ──────────────────────────────────────────────────────────────────
class TestMtimeInvalidation:
    def test_mtime_change_triggers_reread(self, sample_xlsx: Path) -> None:
        """改文件 mtime → key 变 → 不命中缓存，重新读取。

        注意：单纯改 mtime 不改内容也算"变了"——缓存键以 mtime 为准，
        不做内容 hash（hash 几万行太慢，且 mtime 改了就意味着用户重新
        生成了 xlsx，重读更安全）。
        """
        from civ_core.core.data_cache import ExcelDataCache

        cache = ExcelDataCache()
        r1 = cache.get_rows(sample_xlsx, None, 1)

        # 文件系统 mtime 精度通常 1s，往未来推 2s 保证不同
        new_mtime = sample_xlsx.stat().st_mtime + 2
        os.utime(sample_xlsx, (new_mtime, new_mtime))

        r2 = cache.get_rows(sample_xlsx, None, 1)
        # 内容相同（只改了 mtime）但 list 对象应该是新的（不命中缓存）
        assert r1 is not r2
        assert r1 == r2


# ──────────────────────────────────────────────────────────────────
# 管理：clear / invalidate
# ──────────────────────────────────────────────────────────────────
class TestCacheManagement:
    def test_clear_drops_all_entries(self, sample_xlsx: Path) -> None:
        from civ_core.core.data_cache import ExcelDataCache

        cache = ExcelDataCache()
        r1 = cache.get_rows(sample_xlsx, None, 1)
        cache.clear()
        r2 = cache.get_rows(sample_xlsx, None, 1)
        assert r1 is not r2

    def test_global_singleton_persists(self, sample_xlsx: Path) -> None:
        """模块级 EXCEL_DATA_CACHE 单例应可跨 import 复用。"""
        from civ_core.core.data_cache import EXCEL_DATA_CACHE as c1
        from civ_core.core.data_cache import EXCEL_DATA_CACHE as c2

        assert c1 is c2
        r1 = c1.get_rows(sample_xlsx, None, 1)
        # 用第二个引用读，应命中第一个引用塞的缓存
        r2 = c2.get_rows(sample_xlsx, None, 1)
        assert r1 is r2
        c1.clear()  # 清掉避免污染其他测试


# ──────────────────────────────────────────────────────────────────
# 异常透传：底层 ExcelReadError 不被吞
# ──────────────────────────────────────────────────────────────────
class TestErrorPassthrough:
    def test_nonexistent_path_raises(self, tmp_path: Path) -> None:
        from civ_core.core.data_cache import ExcelDataCache
        from civ_core.infra_io.excel_reader import ExcelReadError

        cache = ExcelDataCache()
        with pytest.raises(ExcelReadError):
            cache.get_rows(tmp_path / "nope.xlsx", None, 1)
