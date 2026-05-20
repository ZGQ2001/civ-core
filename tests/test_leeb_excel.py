"""里氏硬度报检单 Excel 导入/导出测试。

用真实的 D 号站房报检单数据 + 临时 xlsx round-trip 验证：
  1. 报检单格式解析（首行带元信息，后 2 行只有 HL）
  2. 计算结果导出回 xlsx，sheet 数与关键单元格正确
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from civ_core.core.calc_functions import (
    calc_leeb_hardness_batch,
    calc_leeb_hardness_workbook,
)
from civ_core.infra_io.leeb_excel import (
    read_leeb_components,
    read_leeb_workbook,
    write_leeb_results,
    write_leeb_results_workbook,
)
from civ_core.infra_io.standards_db import StandardsDB, seed_all_leeb_tables
from civ_core.utils.exceptions import InfraIOError, InputError

REAL_REPORT_XLSX = Path("data/training_materials/防火厚度报检单(D号站房)新.xlsx")


@pytest.fixture
def db() -> StandardsDB:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    d = StandardsDB(conn)
    d.create_tables()
    seed_all_leeb_tables(d)
    return d


# ── 用真实报检单读 ───────────────────────────────────────────────
@pytest.mark.skipif(
    not REAL_REPORT_XLSX.exists(),
    reason="真实报检单文件未到位",
)
def test_read_real_report_steel_columns() -> None:
    """读「里氏硬度（钢柱）」sheet，应得到 N 个构件，每构件 3 测区 × 9 点。"""
    components = read_leeb_components(REAL_REPORT_XLSX, "里氏硬度（钢柱）")
    assert len(components) > 0

    # 抽检第 1 个构件（地上一层2×H钢柱）
    c1 = components[0]
    assert c1.seq == 1
    assert "2×H" in c1.name or "2*H" in c1.name
    assert c1.thickness == 16.0  # 报检单 L 列
    assert len(c1.test_areas_raw) == 3
    assert len(c1.test_areas_raw[0]) == 9
    # 测区 1 第 1 个值
    assert c1.test_areas_raw[0][0] == 467


# ── 端到端：读取 → 计算 → 导出 ──────────────────────────────────
@pytest.mark.skipif(
    not REAL_REPORT_XLSX.exists(),
    reason="真实报检单文件未到位",
)
def test_full_pipeline_with_real_report(
    db: StandardsDB, tmp_path: Path
) -> None:
    """读真实报检单 → 算 → 写新 xlsx，验证导出结构。"""
    components = read_leeb_components(REAL_REPORT_XLSX, "里氏硬度（钢柱）")
    batch = calc_leeb_hardness_batch(components, db=db)

    # 关键不变量：批级特征值在合理范围（Q355 钢抗拉 ≥ 470）
    assert 400 < batch.batch_fb_char_avg < 700

    out = tmp_path / "result.xlsx"
    write_leeb_results(out, batch, angle_degrees=-90.0)
    assert out.exists()

    # 读回验证 2 张 sheet
    wb = load_workbook(str(out), data_only=True)
    assert "原始数据" in wb.sheetnames
    assert "计算结果" in wb.sheetnames

    ws_res = wb["计算结果"]
    # 表头
    assert ws_res.cell(1, 1).value == "序号"
    assert ws_res.cell(1, 8).value == "fb_min(MPa)"
    # 第 1 测区有 HL_m、fb_min
    assert isinstance(ws_res.cell(2, 4).value, int)  # HL_m
    assert isinstance(ws_res.cell(2, 8).value, (int, float))  # fb_min


# ── 异常路径 ─────────────────────────────────────────────────────
def test_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(InfraIOError, match="不存在"):
        read_leeb_components(tmp_path / "nope.xlsx", "Sheet1")


def test_missing_sheet_raises(tmp_path: Path) -> None:
    """造一个空 xlsx，请求不存在的 sheet。"""
    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(str(p))
    with pytest.raises(InputError, match="工作表"):
        read_leeb_components(p, "不存在的sheet")


# ── 合成数据 round-trip（不依赖真实报检单文件）──────────────────
def test_synthetic_round_trip(db: StandardsDB, tmp_path: Path) -> None:
    """造一个仿报检单格式的 xlsx，读 → 算 → 写。"""
    p = tmp_path / "synthetic.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "里氏"
    # 表头
    ws.append(
        ["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9",
         "厚度", "fb", "fb平均", "", "检测批"]
    )
    # 构件 1（3 测区 × 9 点）
    ws.append([1, "Z-1", 467, 465, 471, 468, 467, 468, 473, 472, 463, 16, "", "", "", "批1"])
    ws.append(["", "", 471, 478, 471, 470, 480, 477, 472, 475, 465, "", "", "", "", ""])
    ws.append(["", "", 477, 481, 468, 469, 478, 470, 469, 476, 462, "", "", "", "", ""])
    # 构件 2
    ws.append([2, "Z-2", 470, 472, 471, 469, 473, 475, 470, 472, 471, 16, "", "", "", ""])
    ws.append(["", "", 469, 470, 472, 473, 471, 470, 472, 471, 470, "", "", "", "", ""])
    ws.append(["", "", 471, 472, 470, 471, 470, 473, 469, 471, 472, "", "", "", "", ""])
    wb.save(str(p))

    components = read_leeb_components(p, "里氏")
    assert len(components) == 2
    assert components[0].seq == 1
    assert components[0].thickness == 16.0
    assert components[0].batch_name == "批1"
    assert components[1].seq == 2

    batch = calc_leeb_hardness_batch(components, db=db)
    assert batch.n_components == 2

    out = tmp_path / "out.xlsx"
    write_leeb_results(out, batch, angle_degrees=-90.0)
    wb_out = load_workbook(str(out))
    assert wb_out["计算结果"].max_row >= 8  # 表头 + 6 测区 + 空行 + 批级 = 9


def test_missing_thickness_raises(tmp_path: Path) -> None:
    """厚度列缺失应抛 InputError。"""
    p = tmp_path / "no_thickness.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "里氏"
    ws.append(["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9", "厚度"])
    ws.append([1, "Z-1", 467, 465, 471, 468, 467, 468, 473, 472, 463, None])  # 厚度 None
    wb.save(str(p))
    with pytest.raises(InputError, match="厚度"):
        read_leeb_components(p, "里氏")


def test_missing_hl_value_raises(tmp_path: Path) -> None:
    """HL 9 列缺一应抛 InputError。"""
    p = tmp_path / "no_hl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "里氏"
    ws.append(["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9", "厚度"])
    ws.append([1, "Z-1", 467, 465, 471, None, 467, 468, 473, 472, 463, 12])  # HL4 缺
    wb.save(str(p))
    with pytest.raises(InputError, match="HL"):
        read_leeb_components(p, "里氏")


# ════════════════════════════════════════════════════════════════
# 新格式：read_leeb_workbook 多 sheet 模式
# ════════════════════════════════════════════════════════════════
def _write_new_format_workbook(path: Path) -> None:
    """造一个新格式 xlsx：2 个检测批，每批 2 个构件。"""
    wb = Workbook()
    # 默认 sheet 改名
    ws1 = wb.active
    ws1.title = "检测批1"
    ws1.append(
        ["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9", "厚度"]
    )
    ws1.append([1, "钢柱A-1", 467, 465, 471, 468, 467, 468, 473, 472, 463, 12])
    ws1.append(["", "", 471, 478, 471, 470, 480, 477, 472, 475, 465, ""])
    ws1.append(["", "", 477, 481, 468, 469, 478, 470, 469, 476, 462, ""])
    ws1.append([2, "钢柱A-2", 470, 472, 471, 469, 473, 475, 470, 472, 471, 12])
    ws1.append(["", "", 469, 470, 472, 473, 471, 470, 472, 471, 470, ""])
    ws1.append(["", "", 471, 472, 470, 471, 470, 473, 469, 471, 472, ""])

    ws2 = wb.create_sheet("检测批2")
    ws2.append(
        ["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9", "厚度"]
    )
    ws2.append([1, "钢梁B-1", 460, 462, 461, 459, 463, 465, 460, 462, 461, 10])
    ws2.append(["", "", 459, 460, 462, 463, 461, 460, 462, 461, 460, ""])
    ws2.append(["", "", 461, 462, 460, 461, 460, 463, 459, 461, 462, ""])
    wb.save(str(path))


def test_read_workbook_two_batches(tmp_path: Path) -> None:
    p = tmp_path / "new_format.xlsx"
    _write_new_format_workbook(p)
    wb_in = read_leeb_workbook(p, default_angle_degrees=0.0)
    assert len(wb_in.batches) == 2
    assert wb_in.batches[0].batch_name == "检测批1"
    assert wb_in.batches[1].batch_name == "检测批2"
    assert len(wb_in.batches[0].components) == 2  # 钢柱A-1, A-2
    assert len(wb_in.batches[1].components) == 1  # 钢梁B-1
    # 构件 batch_name 跟随 sheet 名
    assert wb_in.batches[0].components[0].batch_name == "检测批1"


def test_read_workbook_filter_by_sheet_name(tmp_path: Path) -> None:
    """sheet_name_filter='检测批' 过滤掉无关 sheet（如「委托信息」）。"""
    p = tmp_path / "with_meta.xlsx"
    wb = Workbook()
    wb.active.title = "委托信息"
    wb.active.append(["工程名称", "示例项目"])

    ws2 = wb.create_sheet("检测批1")
    ws2.append(
        ["序号", "构件位置", "HL1", "HL2", "HL3", "HL4", "HL5", "HL6", "HL7", "HL8", "HL9", "厚度"]
    )
    ws2.append([1, "C-1", 470, 472, 471, 469, 473, 475, 470, 472, 471, 12])
    ws2.append(["", "", 469, 470, 472, 473, 471, 470, 472, 471, 470, ""])
    ws2.append(["", "", 471, 472, 470, 471, 470, 473, 469, 471, 472, ""])
    wb.save(str(p))

    wb_in = read_leeb_workbook(p, sheet_name_filter="检测批")
    assert len(wb_in.batches) == 1
    assert wb_in.batches[0].batch_name == "检测批1"


def test_full_new_format_pipeline(db: StandardsDB, tmp_path: Path) -> None:
    """读 → 计算 → 导出结果文件，验证每批两 sheet 命名 + 内容。"""
    src = tmp_path / "src.xlsx"
    _write_new_format_workbook(src)
    wb_in = read_leeb_workbook(src, default_angle_degrees=0.0)
    result = calc_leeb_hardness_workbook(wb_in, db=db)

    assert result.n_batches == 2
    assert result.n_components_total == 3
    # batch_name 应跟 sheet 名一致
    assert result.batch_results[0].batch_name == "检测批1"
    assert result.batch_results[1].batch_name == "检测批2"

    # 导出结果文件
    out = tmp_path / "result.xlsx"
    write_leeb_results_workbook(out, result, angle_degrees=0.0)
    wb_out = load_workbook(str(out))
    # 每批 2 sheet → 共 4 sheet
    assert "检测批1-过程数据" in wb_out.sheetnames
    assert "检测批1-报告插入表" in wb_out.sheetnames
    assert "检测批2-过程数据" in wb_out.sheetnames
    assert "检测批2-报告插入表" in wb_out.sheetnames
    assert len(wb_out.sheetnames) == 4

    # 过程 sheet 表头校验
    ws_proc = wb_out["检测批1-过程数据"]
    assert ws_proc.cell(1, 1).value == "序号"
    assert ws_proc.cell(1, 8).value == "fb_min (MPa)"
    # 报告 sheet 表头
    ws_rep = wb_out["检测批1-报告插入表"]
    assert ws_rep.cell(1, 1).value == "检测部位"
